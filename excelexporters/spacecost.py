import base64
import uuid
import os
from openpyxl.chart import (
    PieChart,
    BarChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):
    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 118
    for i in range(2, 37 + 1):
        ws.row_dimensions[i].height = 30

    for i in range(38, 69 + 1):
        ws.row_dimensions[i].height = 30

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('I')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    # Img
    img = Image("excelexporters/myems.png")
    # img = Image("myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws.merge_cells("G3:H3")
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ##################################

    reporting_period_data = report['reporting_period']

    has_cost_data_flag = True

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        has_cost_data_flag = False

    if has_cost_data_flag:
        ws['B6'].font = title_font
        ws['B6'] = name + ' 报告期成本'

        category = reporting_period_data['names']
        ca_len = len(category)

        ws['B7'].fill = table_fill

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = '成本'
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = '单位面积值'
        ws['B9'].border = f_border

        ws['B10'].font = title_font
        ws['B10'].alignment = c_c_alignment
        ws['B10'] = '环比'
        ws['B10'].border = f_border

        for i in range(0, ca_len):
            col = chr(ord('C') + i)

            ws[col + '7'].fill = table_fill
            ws[col + '7'].font = name_font
            ws[col + '7'].alignment = c_c_alignment
            ws[col + '7'] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
            ws[col + '7'].border = f_border

            ws[col + '8'].font = name_font
            ws[col + '8'].alignment = c_c_alignment
            ws[col + '8'] = round(reporting_period_data['subtotals'][i], 0)
            ws[col + '8'].border = f_border

            ws[col + '9'].font = name_font
            ws[col + '9'].alignment = c_c_alignment
            ws[col + '9'] = round(reporting_period_data['subtotals_per_unit_area'][i], 2)
            ws[col + '9'].border = f_border

            ws[col + '10'].font = name_font
            ws[col + '10'].alignment = c_c_alignment
            ws[col + '10'] = str(round(reporting_period_data['increment_rates'][i] * 100, 2)) + "%" \
                if reporting_period_data['increment_rates'][i] is not None else "-"
            ws[col + '10'].border = f_border

    else:
        for i in range(6, 10 + 1):
            ws.row_dimensions[i].height = 0.1

    ##############################

    has_ele_peak_flag = True
    if "toppeaks" not in reporting_period_data.keys() or \
            reporting_period_data['toppeaks'] is None or \
            len(reporting_period_data['toppeaks']) == 0:
        has_ele_peak_flag = False

    if has_ele_peak_flag:
        ws['B12'].font = title_font
        ws['B12'] = name + '分时用电成本'

        ws['B13'].fill = table_fill
        ws['B13'].font = name_font
        ws['B13'].alignment = c_c_alignment
        ws['B13'].border = f_border

        ws['C13'].fill = table_fill
        ws['C13'].font = name_font
        ws['C13'].alignment = c_c_alignment
        ws['C13'].border = f_border
        ws['C13'] = '分时用电成本'

        ws['B14'].font = title_font
        ws['B14'].alignment = c_c_alignment
        ws['B14'] = '尖'
        ws['B14'].border = f_border

        ws['C14'].font = title_font
        ws['C14'].alignment = c_c_alignment
        ws['C14'].border = f_border
        ws['C14'] = round(reporting_period_data['toppeaks'][0], 0)

        ws['B15'].font = title_font
        ws['B15'].alignment = c_c_alignment
        ws['B15'] = '峰'
        ws['B15'].border = f_border

        ws['C15'].font = title_font
        ws['C15'].alignment = c_c_alignment
        ws['C15'].border = f_border
        ws['C15'] = round(reporting_period_data['onpeaks'][0], 0)

        ws['B16'].font = title_font
        ws['B16'].alignment = c_c_alignment
        ws['B16'] = '平'
        ws['B16'].border = f_border

        ws['C16'].font = title_font
        ws['C16'].alignment = c_c_alignment
        ws['C16'].border = f_border
        ws['C16'] = round(reporting_period_data['midpeaks'][0], 0)

        ws['B17'].font = title_font
        ws['B17'].alignment = c_c_alignment
        ws['B17'] = '谷'
        ws['B17'].border = f_border

        ws['C17'].font = title_font
        ws['C17'].alignment = c_c_alignment
        ws['C17'].border = f_border
        ws['C17'] = round(reporting_period_data['offpeaks'][0], 0)

        pie = PieChart()
        labels = Reference(ws, min_col=2, min_row=14, max_row=17)
        pie_data = Reference(ws, min_col=3, min_row=13, max_row=17)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 5.25
        pie.width = 9
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True

        ws.add_chart(pie, "D13")

    else:
        for i in range(12, 18 + 1):
            ws.row_dimensions[i].height = 0.1

    ##################################
    current_row_flag = 19

    has_child_flag = True
    if "child_space" not in report.keys() or "energy_category_names" not in report['child_space'].keys() or \
            len(report['child_space']["energy_category_names"]) == 0:
        has_child_flag = False

    if has_child_flag:
        child = report['child_space']

        ws['B' + str(current_row_flag)].font = title_font
        ws['B' + str(current_row_flag)] = name + ' 子空间数据'

        current_row_flag += 1

        ws['B' + str(current_row_flag)].fill = table_fill
        ws['B' + str(current_row_flag)].border = f_border
        ca_len = len(child['energy_category_names'])

        for i in range(0, ca_len):
            row = chr(ord('C') + i)
            ws[row + str(current_row_flag)].fill = table_fill
            ws[row + str(current_row_flag)].font = title_font
            ws[row + str(current_row_flag)].alignment = c_c_alignment
            ws[row + str(current_row_flag)].border = f_border
            ws[row + str(current_row_flag)] = child['energy_category_names'][i]

        space_len = len(child['child_space_names_array'][0])

        for i in range(0, space_len):
            current_row_flag += 1
            row = str(current_row_flag)

            ws['B' + row].font = name_font
            ws['B' + row].alignment = c_c_alignment
            ws['B' + row] = child['child_space_names_array'][0][i]
            ws['B' + row].border = f_border

            for j in range(0, ca_len):
                col = chr(ord('C') + j)
                ws[col + row].font = name_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = child['subtotals_array'][j][i]
                ws[col + row].border = f_border

        current_row_flag += 1

        # Pie
        for i in range(0, ca_len):
            pie = PieChart()
            labels = Reference(ws, min_col=2, min_row=current_row_flag - space_len, max_row=current_row_flag - 1)
            pie_data = Reference(ws, min_col=3 + i, min_row=current_row_flag - space_len - 1,
                                 max_row=current_row_flag - 1)
            pie.add_data(pie_data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 5.25
            pie.width = 8
            col = chr(ord('C') + i)
            pie.title = ws[col + '20'].value
            s1 = pie.series[0]
            s1.dLbls = DataLabelList()
            s1.dLbls.showCatName = False
            s1.dLbls.showVal = True
            s1.dLbls.showPercent = True
            chart_col = chr(ord('B') + 2 * i)
            chart_cell = chart_col + str(current_row_flag)
            ws.add_chart(pie, chart_cell)

        current_row_flag += 5

    else:
        for i in range(19, 36 + 1):
            current_row_flag = 36
            ws.row_dimensions[i].height = 0.1

    current_row_flag += 1

    #############################################
    reporting_period_data = report['reporting_period']
    times = reporting_period_data['timestamps']
    has_detail_data_flag = True
    ca_len = len(report['reporting_period']['names'])
    table_row = (current_row_flag + 1) + ca_len * 5
    if "timestamps" not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0:
        has_detail_data_flag = False

    if has_detail_data_flag:
        ws['B' + str(current_row_flag)].font = title_font
        ws['B' + str(current_row_flag)] = name + ' 详细数据'

        ws['B' + str(table_row)].fill = table_fill
        ws['B' + str(table_row)].border = f_border
        ws['B' + str(table_row)].alignment = c_c_alignment
        ws['B' + str(table_row)] = '日期时间'
        time = times[0]
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = table_row + len(time)

        if has_data:
            for i in range(0, len(time)):
                col = 'B'
                row = str(table_row + 1 + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):

                col = chr(ord('C') + i)

                ws[col + str(table_row)].fill = table_fill
                ws[col + str(table_row)].font = title_font
                ws[col + str(table_row)].alignment = c_c_alignment
                ws[col + str(table_row)] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][
                    i] + ")"
                ws[col + str(table_row)].border = f_border

                # 39 data
                time = times[i]
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(table_row + 1 + j)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(reporting_period_data['values'][i][j], 0)
                    ws[col + row].border = f_border

                bar = BarChart()
                labels = Reference(ws, min_col=2, min_row=table_row + 1, max_row=max_row)
                bar_data = Reference(ws, min_col=3 + i, min_row=table_row, max_row=max_row)
                bar.add_data(bar_data, titles_from_data=True)
                bar.set_categories(labels)
                bar.height = 5.25
                bar.width = 18
                bar.dLbls = DataLabelList()
                bar.dLbls.showVal = True
                bar.dLbls.showPercent = True
                chart_col = 'B'
                chart_cell = chart_col + str(current_row_flag + 1 + 5 * i)
                ws.add_chart(bar, chart_cell)
    else:
        for i in range(37, 69 + 1):
            ws.row_dimensions[i].height = 0.1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
